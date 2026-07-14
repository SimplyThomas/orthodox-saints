#!/usr/bin/env python3
"""Build tool for the Orthodox Saints Database.

Pipeline:  data/*.csv  ->  in-memory SQLite  ->  validate  ->  emit artifacts.

The CSVs in data/ are the SOURCE OF TRUTH. SQLite is created fresh every run,
used only for validation/query, then discarded. Generated output (public/, dist/)
is never committed. See CLAUDE.md and docs/historical/bootstrap.md.

Usage:
    python build.py                 validate + emit data.json, xlsx
    python build.py --check-only    validate only, exit non-zero on any violation
    python build.py --xlsx-only     emit only the Excel export
    python build.py --sqlite        also emit public/saints.sqlite (read-only artifact)
    python build.py --report        rank icon-less saints by priority (authoring aid)

Map of this file (sections in execution order, each under a dashed banner):
    load_*()            read the CSVs (saints, vocab, images, quotes, groups, ...)
    assign_ids()        fill blank Saint IDs with the next OS-#### and
                        write_saints() the CSV back — the one place source is mutated
    parse_* / feast_*   feast-string helpers (months, sort key, day-range checks)
    build_db()          throwaway in-memory SQLite used for validation queries
    validate()          collect EVERY violation (never stop at the first), plus
                        warnings; main() exits non-zero if any error
    report_coverage() / authoring reports printed to stdout (finder coverage,
    report_priority()   icon-priority ranking) — no files written
    derive_links()      Google/YouTube search-URL columns (18/19/25)
    to_record()         one CSV row -> one JSON record (joins images, quotes,
                        groups, themes, name variants, search haystack)
    emit_*()            write public/data.json, groups.json, themes.json,
                        feasts.json (via feastlib), saints.sqlite, dist/*.xlsx
Feasts & Fasts (data/feasts.csv) is loaded/validated/emitted by feastlib.py,
orchestrated from main().
"""

from __future__ import annotations

import argparse
import csv
import json
import os
import re
import sqlite3
import sys
import unicodedata
import urllib.parse
from pathlib import Path

import feastlib
import hostlib
import themes as themes_mod

ROOT = Path(__file__).resolve().parent
DATA = ROOT / "data"
SRC = ROOT / "src"
PROFILES_DIR = SRC / "content" / "profiles"   # per-saint rich profiles (YAML)
PUBLIC = ROOT / "public"
DIST = ROOT / "dist"
STATIC = ROOT / "static"  # Astro publicDir; self-hosted icons live in static/icons/

SAINTS_CSV = DATA / "saints.csv"
VOCAB_CSV = DATA / "vocabulary.csv"
VENDORS_CSV = DATA / "vendors.csv"
NAME_VARIANTS_CSV = DATA / "name_variants.csv"
SAINT_IMAGES_CSV = DATA / "saint_images.csv"
SAINT_QUOTES_CSV = DATA / "saint_quotes.csv"
GROUPS_CSV = DATA / "groups.csv"
SAINT_GROUPS_CSV = DATA / "saint_groups.csv"
RETIRED_IDS_CSV = DATA / "retired_ids.csv"
RETIRED_IDS_HEADER = ["retired_id", "retired_name", "canonical_id", "canonical_name",
                      "reason", "date", "pr"]

# Group taxonomy (data/groups.csv + data/saint_groups.csv) — a first-class way to
# re-link members of a collective commemoration (a synaxis, feast-companions, a
# household). Two join files following the saint_images/saint_quotes pattern;
# referential integrity is enforced at build time (fail loud). `type` is a small
# enumerated set — adding one is a deliberate code change.
GROUPS_HEADER = ["slug", "saint_id", "name", "type", "description", "feast",
                 "sort", "rule"]
SAINT_GROUPS_HEADER = ["group_slug", "saint_id", "role", "order"]
GROUP_TYPES = {"synaxis", "feast-companions", "household"}
SLUG_RE = re.compile(r"^[a-z0-9]+(?:-[a-z0-9]+)*$")

# Real saint portraits (data/saint_images.csv) join to saints by Saint ID. Only
# OPEN, reusable licenses are accepted — an unlicensed image must never deploy
# (CLAUDE.md §9). Attribution licenses (CC-BY*) additionally require a credit.
SAINT_IMAGES_HEADER = ["saint_id", "image_path", "license", "credit", "source"]
OPEN_LICENSES = {"PD", "PD-art", "PD-old", "CC0"}  # public-domain / no-rights
# The accepted-license list as shown in error messages (tests assert on it).
OPEN_LICENSE_LIST = "PD / PD-art / PD-old / CC0 / CC-BY / CC-BY-SA"

# Vendor-permission image registry (data/image_permissions.csv). A "used with
# permission" image is NOT an open license — it is a revocable, per-vendor grant
# (CLAUDE.md §9). Each saint_images.csv row that uses one carries a license token
# `Permission:<vendor_slug>` joined to a row here. `status` is the kill-switch:
# flip to `revoked` and the build stops publishing that vendor's images.
IMAGE_PERMISSIONS_CSV = DATA / "image_permissions.csv"
IMAGE_PERMISSIONS_HEADER = [
    "vendor_slug", "vendor_name", "attribution", "homepage", "granted", "status", "terms"
]
PERMISSION_STATUSES = {"active", "revoked"}
PERMISSION_LICENSE_RE = re.compile(r"^Permission:([a-z0-9]+(?:-[a-z0-9]+)*)$")

# Additional depictions (data/saint_depictions.csv) power the saint page's
# "Depictions & Icons" carousel — MANY images per saint (museum icons, PD masters,
# and vendor-permission icons available to commission or order). Same licensing
# gate as saint_images (an open license OR a Permission:<vendor> token), but
# multiple rows per saint, each carrying the card-presentation columns
# (kind/tag/title/era/by). `kind` drives the card tone; row order is carousel order.
SAINT_DEPICTIONS_CSV = DATA / "saint_depictions.csv"
SAINT_DEPICTIONS_HEADER = ["saint_id", "image_path", "license", "credit", "source",
                           "kind", "tag", "title", "era", "by"]
DEPICTION_KINDS = {"museum", "iconographer", "shop"}


def permission_slug(lic: str) -> str | None:
    """Return the vendor slug if `lic` is a `Permission:<slug>` token, else None."""
    m = PERMISSION_LICENSE_RE.match(lic.strip())
    return m.group(1) if m else None


def license_ok(lic: str) -> bool:
    """True if the license is an accepted open license (public-domain family or
    any Creative Commons Attribution variant: CC-BY / CC-BY-SA)."""
    lic = lic.strip()
    return lic in OPEN_LICENSES or bool(re.match(r"^CC-BY(-SA)?(-\d(\.\d)?)?$", lic))


def license_requires_credit(lic: str) -> bool:
    return lic.strip().upper().startswith("CC-BY")


def validate_image_license(where: str, sid: str, lic: str, credit: str,
                           source: str, permissions: dict[str, dict[str, str]],
                           *, subject: str, noun: str
                           ) -> tuple[list[str], list[str]]:
    """The licensing gate shared by saint_images and saint_depictions rows
    (CLAUDE.md §9): an accepted open license (with a credit when required) OR
    a Permission:<vendor> token validated against the permission registry —
    a revoked vendor warns (the row is excluded from output, not an error).
    `subject` ("Self-hosted images" / "Depictions") and `noun` ("image" /
    "depiction") only vary the message wording."""
    errors: list[str] = []
    warnings: list[str] = []
    slug = permission_slug(lic)
    if not lic:
        errors.append(f"{where} ({sid}): empty license. {subject} must "
                      "declare an open license or a Permission:<vendor> "
                      "token (§9).")
    elif slug is not None:
        vendor = permissions.get(slug)
        if vendor is None:
            errors.append(f"{where} ({sid}): permission vendor {slug!r} is "
                          "not in data/image_permissions.csv.")
        elif vendor.get("status") == "revoked":
            warnings.append(f"{where} ({sid}): vendor {slug!r} permission is "
                            f"REVOKED — {noun} excluded from output; delete "
                            f"the file under static/icons/permission/{slug}/.")
        elif not source:
            errors.append(f"{where} ({sid}): permission {noun} requires a "
                          "'source' linking the specific vendor icon page (§9).")
    elif not license_ok(lic):
        errors.append(f"{where} ({sid}): license {lic!r} is not an accepted "
                      f"open license ({OPEN_LICENSE_LIST}) or a "
                      "Permission:<vendor> token.")
    elif license_requires_credit(lic) and not credit:
        errors.append(f"{where} ({sid}): license {lic} requires a 'credit' "
                      "(attribution).")

    if not source and slug is None:
        warnings.append(f"{where} ({sid}): no 'source' (provenance) given.")
    return errors, warnings


# Saint quotes (data/saint_quotes.csv) join to saints by Saint ID. Each quote
# MUST come from a public-domain translation — the saint's original words are
# PD, but a modern English translation usually is not, and §9 forbids reproducing
# a copyrighted translation. The translation field is gated like the image
# license: it must name a PD source (ANF / NPNF / explicit (PD) / CC0); anything
# else fails the build. One quote per saint (detail page shows a single quote).
SAINT_QUOTES_HEADER = ["saint_id", "quote", "work", "locus", "translation", "source_url"]
PD_TRANSLATION_RE = re.compile(r"\b(PD|PD-old|CC0|ANF|NPNF1?|NPNF2)\b", re.IGNORECASE)


def translation_ok(t: str) -> bool:
    """True if the translation names an accepted public-domain source — the
    Ante-/Nicene-and-Post-Nicene-Fathers series (ANF / NPNF / NPNF1 / NPNF2),
    an explicit public-domain marker (PD / PD-old), or CC0. A modern in-copyright
    translation (e.g. a Philokalia or SVS Press edition) has no such marker and
    fails, keeping copyrighted translations out of the data (§9)."""
    return bool(PD_TRANSLATION_RE.search(t or ""))

# The canonical 26-column header, exact and in order (CLAUDE.md §5).
HEADER = [
    "Saint ID", "Name", "Also Known As", "Gender", "Rank / Type", "Church Status",
    "Family / Life State", "Vocation", "Life Experience", "Virtue",
    "Commonly Asked Intercessions", "Region of Origin", "Tradition of Veneration",
    "Era", "Century", "Feast Day(s)", "Short Prayer (Intercession)",
    "Hymn / Apolytikion", "Icon", "Brief Life", "Notes", "Customs & Traditions",
    "Works by the Saint", "Works About the Saint", "Video / Media", "Sources",
]

# Controlled columns -> their vocabulary category (same name).
CONTROLLED = [
    "Gender", "Rank / Type", "Church Status", "Family / Life State", "Vocation",
    "Life Experience", "Virtue", "Commonly Asked Intercessions", "Region of Origin",
    "Tradition of Veneration", "Era", "Century",
]
# Single-value controlled columns (the rest are multi-value).
SINGLE_VALUE = {"Gender", "Era", "Century"}

# Required non-empty fields (bootstrap §4). Era-or-Century handled separately.
# Feast Day(s) is NOT required: a handful of genuinely-commemorated saints have no
# fixed (or only a movable/undocumented) feast and are entered as stubs without one.
# When present it must still parse (see feast_recognized check below).
REQUIRED = ["Name", "Rank / Type", "Gender",
            "Short Prayer (Intercession)", "Sources"]

# Free-text multi-value columns split for the JSON (not vocab-checked).
FREE_MULTI = ["Also Known As", "Works by the Saint", "Works About the Saint"]

MULTI_SEP = "; "

MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
          "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
MONTH_INDEX = {m: i + 1 for i, m in enumerate(MONTHS)}
FEAST_RE = re.compile(r"\b(" + "|".join(MONTHS) + r")\s+(\d{1,2})\b")

# Movable feasts are a permanent feature of the Orthodox calendar (Pascha- and
# Pentecost-relative). They have no fixed "Mon D" token; we accept them as valid
# feasts and sort them after the fixed-date calendar.
MOVABLE_RE = re.compile(
    r"\b(Sunday|Saturday|Pascha|Pentecost|Lent|Lenten|Cheesefare|Meatfare|"
    r"Pentecostarion|Ascension|Thomas|Palm|Holy Week|Bright)\b",
    re.IGNORECASE,
)
MOVABLE_SORT = 9999  # sorts movable-only feasts last in calendar order

ID_RE = re.compile(r"^OS-\d{4,}$")

# Short JSON keys, stable, in display order. Maps CSV column -> json key.
JSON_KEYS = {
    "Saint ID": "id", "Name": "name", "Also Known As": "aka", "Gender": "gender",
    "Rank / Type": "rank", "Church Status": "church", "Family / Life State": "family",
    "Vocation": "vocation", "Life Experience": "experience", "Virtue": "virtue",
    "Commonly Asked Intercessions": "intercession", "Region of Origin": "origin",
    "Tradition of Veneration": "tradition", "Era": "era", "Century": "century",
    "Feast Day(s)": "feast", "Short Prayer (Intercession)": "prayer",
    "Hymn / Apolytikion": "hymn", "Icon": "icon", "Brief Life": "brief",
    "Notes": "notes", "Customs & Traditions": "customs",
    "Works by the Saint": "works", "Works About the Saint": "about",
    "Video / Media": "video", "Sources": "sources",
}
# Which json keys are arrays (multi-value).
ARRAY_KEYS = {JSON_KEYS[c] for c in (
    [c for c in CONTROLLED if c not in SINGLE_VALUE] + FREE_MULTI
)}


def split_multi(value: str) -> list[str]:
    return [v.strip() for v in value.split(MULTI_SEP) if v.strip()]


# --------------------------------------------------------------------------- #
# Load
# --------------------------------------------------------------------------- #
def load_vocab() -> dict[str, set[str]]:
    vocab: dict[str, set[str]] = {}
    with open(VOCAB_CSV, encoding="utf-8", newline="") as f:
        reader = csv.reader(f)
        header = next(reader, None)
        if header != ["category", "term"]:
            sys.exit(f"FATAL: {VOCAB_CSV} header must be 'category,term', got {header!r}")
        for row in reader:
            if not row or not any(c.strip() for c in row):
                continue
            cat, term = row[0].strip(), row[1].strip()
            vocab.setdefault(cat, set()).add(term)
    return vocab


def load_saints() -> tuple[list[str], list[dict[str, str]]]:
    with open(SAINTS_CSV, encoding="utf-8", newline="") as f:
        reader = csv.reader(f)
        header = next(reader, None)
        rows = [dict(zip(header, r)) for r in reader if any(c.strip() for c in r)]
    return header, rows


# --------------------------------------------------------------------------- #
# ID assignment (CLAUDE.md §6): blank IDs get the next sequential OS-####.
# --------------------------------------------------------------------------- #
def next_id_seed(rows: list[dict[str, str]], id_field: str = "Saint ID") -> int:
    """Highest existing OS-#### number across rows (0 if none)."""
    max_num = 0
    for r in rows:
        m = re.match(r"^OS-(\d+)$", (r.get(id_field) or "").strip())
        if m:
            max_num = max(max_num, int(m.group(1)))
    return max_num


def retired_id_seed() -> int:
    """Highest OS-#### number in retired_ids.csv (0 if none/absent). Retired IDs
    are permanently spent (§6) — the counter must clear them so a new row never
    reuses one, even when the retired ID exceeds every active ID."""
    if not RETIRED_IDS_CSV.exists():
        return 0
    max_num = 0
    with RETIRED_IDS_CSV.open(encoding="utf-8", newline="") as f:
        for row in csv.DictReader(f):
            m = re.match(r"^OS-(\d+)$", (row.get("retired_id") or "").strip())
            if m:
                max_num = max(max_num, int(m.group(1)))
    return max_num


def assign_ids(rows: list[dict[str, str]], id_field: str = "Saint ID",
               seed: int | None = None) -> tuple[bool, int]:
    """Assign the next sequential OS-#### to any blank `id_field`. Mutates rows
    in place; returns (any_assigned, highest_number_now_used). Pure (no file
    I/O). `seed` lets callers share ONE counter across files (saints + groups)
    so the two id spaces never collide (§6: opaque, permanent, never reused)."""
    max_num = next_id_seed(rows, id_field) if seed is None else seed
    assigned = False
    for r in rows:
        if not (r.get(id_field) or "").strip():
            max_num += 1
            r[id_field] = f"OS-{max_num:04d}"
            assigned = True
            label = r.get("Name") or r.get("name") or r.get("slug") or ""
            print(f"  assigned {r[id_field]}  {label}")
    return assigned, max_num


def write_saints(header: list[str], rows: list[dict[str, str]]) -> None:
    """Rewrite data/saints.csv in place (after assign_ids fills blank IDs).

    The ONLY function that mutates a source-of-truth file. newline="" hands
    line-ending control to csv.writer, which emits \\r\\n — preserving the
    file's CRLF convention (CLAUDE.md §7)."""
    with open(SAINTS_CSV, "w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=header)
        w.writeheader()
        w.writerows(rows)
    print(f"  wrote stable IDs back to {SAINTS_CSV.relative_to(ROOT)}")


def load_group_rows() -> tuple[list[str] | None, list[dict[str, str]]]:
    """Raw groups.csv rows keyed by column name, normalized to GROUPS_HEADER
    (so a pre-saint_id file is migrated on the next write-back). Returns
    (None, []) when the file is absent."""
    if not GROUPS_CSV.exists():
        return None, []
    with open(GROUPS_CSV, encoding="utf-8", newline="") as f:
        reader = csv.reader(f)
        header = next(reader, None) or []
        rows = []
        for r in reader:
            if not any(c.strip() for c in r):
                continue
            row = dict(zip(header, r))
            rows.append({k: row.get(k, "") for k in GROUPS_HEADER})
    return GROUPS_HEADER, rows


def write_groups(header: list[str], rows: list[dict[str, str]]) -> None:
    """Rewrite data/groups.csv in place after assign_ids fills blank saint_ids.
    Like the other data CSVs, groups.csv is CRLF (CLAUDE.md §7; crlf_errors
    enforces it), so pin the lineterminator to keep the diff to the id column."""
    with open(GROUPS_CSV, "w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=header, lineterminator="\r\n")
        w.writeheader()
        w.writerows(rows)
    print(f"  wrote stable group IDs back to {GROUPS_CSV.relative_to(ROOT)}")


# --------------------------------------------------------------------------- #
# Feast parsing
# --------------------------------------------------------------------------- #
def parse_months(feast: str) -> list[str]:
    seen: list[str] = []
    for mon, _day in FEAST_RE.findall(feast):
        if mon not in seen:
            seen.append(mon)
    return seen


def feast_sort(feast: str) -> int:
    vals = [MONTH_INDEX[mon] * 100 + int(day) for mon, day in FEAST_RE.findall(feast)]
    if vals:
        return min(vals)
    return MOVABLE_SORT  # movable-only feast


def feast_recognized(feast: str) -> bool:
    """A feast is valid if it has a fixed Mon-D token OR a movable-feast keyword."""
    return bool(FEAST_RE.search(feast) or MOVABLE_RE.search(feast))


# Max day per month. Feb 29 is allowed: the fixed calendar has Feb 29
# commemorations (kept on Feb 28 in non-leap years).
MONTH_MAX_DAY = {"Jan": 31, "Feb": 29, "Mar": 31, "Apr": 30, "May": 31,
                 "Jun": 30, "Jul": 31, "Aug": 31, "Sep": 30, "Oct": 31,
                 "Nov": 30, "Dec": 31}


def feast_day_range_errors(feast: str) -> list[str]:
    """Mon-D tokens whose day falls outside the month's real range.

    FEAST_RE accepts any 1-2 digit day, so 'Feb 30' or 'Sep 0' would
    otherwise ship silently (and sort to a nonsense feast_sort)."""
    return [f"{mon} {day}" for mon, day in FEAST_RE.findall(feast)
            if not 1 <= int(day) <= MONTH_MAX_DAY[mon]]


# --------------------------------------------------------------------------- #
# SQLite build (validation/query engine; discarded unless --sqlite)
# --------------------------------------------------------------------------- #
def build_db(header: list[str], rows: list[dict[str, str]],
             vocab: dict[str, set[str]]) -> sqlite3.Connection:
    """Load saints + vocabulary into a fresh in-memory SQLite database.

    All columns TEXT, plus two derived: `months` (distinct feast months) and
    `feast_sort` (earliest fixed date as MM*100+DD; movable-only sorts last).
    Used for validation/ad-hoc queries and discarded — persisted only when
    --sqlite emits it as a read-only artifact. Never a source of truth."""
    conn = sqlite3.connect(":memory:")
    cur = conn.cursor()
    cols = [f'"{c}"' for c in header] + ["months", "feast_sort"]
    cur.execute(f"CREATE TABLE saints ({', '.join(c + ' TEXT' for c in cols)})")
    placeholders = ", ".join("?" for _ in cols)
    for r in rows:
        values = [r[c] for c in header]
        values.append(MULTI_SEP.join(parse_months(r["Feast Day(s)"])))
        values.append(str(feast_sort(r["Feast Day(s)"])))
        cur.execute(f"INSERT INTO saints VALUES ({placeholders})", values)
    cur.execute("CREATE TABLE vocabulary (category TEXT, term TEXT)")
    for cat, terms in vocab.items():
        cur.executemany("INSERT INTO vocabulary VALUES (?, ?)",
                        [(cat, t) for t in sorted(terms)])
    conn.commit()
    return conn


def crlf_errors(path: Path) -> list[str]:
    """The data CSVs are CRLF (CLAUDE.md §7). An editor that normalizes to LF
    silently corrupts them; catch it at validate time instead of in review.
    Invariant: every \\n is part of a \\r\\n pair."""
    data = path.read_bytes()
    if data.count(b"\n") != data.count(b"\r\n"):
        return [f"{path.name}: contains bare LF line endings — the data CSVs "
                f"are CRLF. Fix your editor/git config (git config "
                f"core.autocrlf false) and restore CRLF before committing."]
    return []


# --------------------------------------------------------------------------- #
# Validate (collect ALL violations, then report)
# --------------------------------------------------------------------------- #
def validate(header: list[str], rows: list[dict[str, str]],
             vocab: dict[str, set[str]]) -> tuple[list[str], list[str]]:
    """Check every invariant and return (errors, warnings) — all of them.

    Never stops at the first violation, so one run surfaces the full fix list.
    Errors fail the build (ID format/uniqueness, required fields, vocab terms,
    feast parsing, image licenses, group/quote/profile referential integrity);
    warnings are advisory (possible duplicate names, missing thumbs/coverage).
    Delegates per-file checks to the validate_* helpers below."""
    errors: list[str] = []
    warnings: list[str] = []

    errors.extend(validate_name_variants())

    for _csv in sorted(DATA.glob("*.csv")):
        errors.extend(crlf_errors(_csv))

    # Always validate saint_images.csv against the committed saints.csv, not just
    # the rows under test — ensures IDs in image rows resolve to real saints even
    # when validate() is called with a synthetic (unit-test) subset.
    try:
        _, _all_saints = load_saints()
        _img_valid_ids = {r["Saint ID"].strip() for r in _all_saints
                         if r["Saint ID"].strip()}
    except Exception:
        _img_valid_ids = {r["Saint ID"].strip() for r in rows if r["Saint ID"].strip()}
    permissions = load_image_permissions()
    perm_errors, perm_warnings = validate_image_permissions()
    errors.extend(perm_errors)
    warnings.extend(perm_warnings)
    img_errors, img_warnings = validate_saint_images(_img_valid_ids, permissions)
    errors.extend(img_errors)
    warnings.extend(img_warnings)

    dep_errors, dep_warnings = validate_saint_depictions(_img_valid_ids, permissions)
    errors.extend(dep_errors)
    warnings.extend(dep_warnings)

    # Validate saint_quotes.csv against the full committed saints.csv too, for the
    # same reason as images (IDs must resolve even under a unit-test subset).
    quote_errors, quote_warnings = validate_saint_quotes(_img_valid_ids)
    errors.extend(quote_errors)
    warnings.extend(quote_warnings)

    # Group profiles (a synaxis / household / …) are also served at /saint/[id]
    # and carry a rich OS-####.yaml, but their IDs live in groups.csv, not
    # saints.csv — so admit them to the set the profile cross-check validates
    # against.
    try:
        _group_ids = {g["saint_id"].strip() for g in load_groups()
                      if g.get("saint_id", "").strip()}
    except Exception:
        _group_ids = set()
    prof_errors, prof_warnings = validate_saint_profiles(_img_valid_ids | _group_ids)
    errors.extend(prof_errors)
    warnings.extend(prof_warnings)

    group_errors, group_warnings = validate_groups(_img_valid_ids)
    errors.extend(group_errors)
    warnings.extend(group_warnings)

    # A deduplicated row may be merged into a group (a synaxis modeled as a
    # group), so a group's OS-#### is a valid canonical target too.
    ret_errors, ret_warnings = validate_retired_ids(_img_valid_ids | _group_ids)
    errors.extend(ret_errors)
    warnings.extend(ret_warnings)

    if header != HEADER:
        errors.append(
            "Header/column mismatch against the canonical 26-column header.\n"
            f"    expected: {HEADER}\n    found:    {header}"
        )
        # Without a matching header, per-row checks are unreliable.
        return errors, warnings

    seen_ids: dict[str, str] = {}
    norm_names: dict[str, list[str]] = {}

    # term -> the categories where it IS a valid vocabulary term. Used to turn a
    # "wrong column" mistake into an actionable hint (the most common authoring slip).
    term_categories: dict[str, list[str]] = {}
    for cat, terms in vocab.items():
        for t in terms:
            term_categories.setdefault(t, []).append(cat)

    for r in rows:
        sid = r["Saint ID"].strip()
        tag = sid or f"(blank id, Name={r['Name']!r})"

        # ID format + uniqueness
        if not sid:
            errors.append(f"{tag}: empty Saint ID (build should have assigned one).")
        elif not ID_RE.match(sid):
            errors.append(f"{sid}: Saint ID does not match ^OS-\\d{{4,}}$.")
        if sid:
            if sid in seen_ids:
                errors.append(f"{sid}: duplicate Saint ID (also '{seen_ids[sid]}').")
            else:
                seen_ids[sid] = r["Name"]

        # Required fields
        for col in REQUIRED:
            if not r[col].strip():
                errors.append(f"{tag}: missing required field '{col}'.")
        if not r["Era"].strip() and not r["Century"].strip():
            errors.append(f"{tag}: requires at least one of Era or Century.")

        # Feast must be parseable (fixed Mon-D or movable keyword)
        if r["Feast Day(s)"].strip() and not feast_recognized(r["Feast Day(s)"]):
            errors.append(
                f"{tag}: unparseable Feast Day(s) {r['Feast Day(s)']!r} "
                "(need a 'Mon D' date or a recognized movable-feast term)."
            )
        for bad in feast_day_range_errors(r["Feast Day(s)"]):
            errors.append(
                f"{tag}: impossible feast date {bad!r} (day out of range "
                "for the month)."
            )

        # Controlled-vocabulary terms
        for col in CONTROLLED:
            raw = r[col].strip()
            if not raw:
                continue
            values = [raw] if col in SINGLE_VALUE else split_multi(raw)
            for v in values:
                if v not in vocab.get(col, set()):
                    other = [c for c in term_categories.get(v, []) if c != col]
                    hint = (f" (it IS a valid '{' / '.join(sorted(other))}' term "
                            "— wrong column?)") if other else ""
                    errors.append(f"{tag}: unknown term in '{col}': {v!r}.{hint}")

        # Warning: near-duplicate name
        key = re.sub(r"[^a-z0-9]", "", r["Name"].lower())
        norm_names.setdefault(key, []).append((sid or r["Name"], r["Notes"]))

        # Warning: finder-coverage nudge on non-stub saints
        is_stub = not r["Brief Life"].strip()
        if not is_stub and not r["Commonly Asked Intercessions"].strip():
            warnings.append(f"{tag}: no Commonly Asked Intercessions (finder coverage).")

        # Validate optional Themes override column
        for slug in themes_mod._split(r.get("Themes", "")):
            if slug not in themes_mod.THEME_SLUGS:
                errors.append(
                    f"{tag}: unknown theme slug {slug!r} in Themes "
                    f"(valid slugs are defined in themes.py)"
                )

    for key, group in norm_names.items():
        if len(group) <= 1:
            continue
        # Documented-distinct: same-name saints verified as different people
        # carry a Notes cross-reference to the other row's ID (e.g. "Distinct
        # from ... (OS-0966)"). Suppress the warning when every member of the
        # group is tied to another member that way, so the warning list stays
        # a real to-investigate queue.
        ids = [sid for sid, _ in group]
        notes = {sid: note for sid, note in group}
        documented = all(
            any(other in notes[sid] or sid in notes[other]
                for other in ids if other != sid)
            for sid in ids
        )
        if not documented:
            warnings.append(f"possible duplicate saint (same normalized name): {ids}")

    return errors, warnings


# --------------------------------------------------------------------------- #
# Finder-coverage report (CLAUDE.md §10: the finder's quality axis)
# --------------------------------------------------------------------------- #
# Facets that drive the finder; reported as fill % so changes are visible.
COVERAGE_COLUMNS = [
    "Commonly Asked Intercessions",
    "Life Experience",
    "Vocation",
    "Virtue",
    "Brief Life",
    "Customs & Traditions",
]


def coverage_stats(rows: list[dict[str, str]]) -> list[tuple[str, int, int, float]]:
    total = len(rows)
    stats = []
    for col in COVERAGE_COLUMNS:
        filled = sum(1 for r in rows if r[col].strip())
        pct = (100.0 * filled / total) if total else 0.0
        stats.append((col, filled, total, pct))
    return stats


def report_coverage(rows: list[dict[str, str]]) -> None:
    """Print finder-coverage stats; also write a Markdown table to the GitHub
    Actions job summary when running in CI ($GITHUB_STEP_SUMMARY)."""
    stats = coverage_stats(rows)
    print("Finder coverage:")
    for col, filled, total, pct in stats:
        print(f"  {col:32s} {filled:4d}/{total}  ({pct:5.1f}%)")

    summary = os.environ.get("GITHUB_STEP_SUMMARY")
    if summary:
        lines = [f"## Finder coverage ({len(rows)} saints)", "",
                 "| Facet | Filled | Coverage |", "|---|---:|---:|"]
        for col, filled, total, pct in stats:
            lines.append(f"| {col} | {filled}/{total} | {pct:.1f}% |")
        with open(summary, "a", encoding="utf-8") as f:
            f.write("\n".join(lines) + "\n")


# --------------------------------------------------------------------------- #
# Priority report (CLAUDE.md §5/§10): rank icon-less saints for the next batch.
# A data-derived proxy for "importance" so each portrait batch is self-directing
# instead of hand-picked. Local authoring aid only — never a CI gate, no output
# file. See issue #83.
# --------------------------------------------------------------------------- #
# Weighted sum of facets all derivable from data/saints.csv. Weights are a
# starting point (issue #83): tradition breadth and a filled Intercession facet
# (the finder's engine, §10) dominate; secondary finder signals and calendar
# surface area break ties.
def priority_score(r: dict[str, str]) -> tuple[int, dict[str, object]]:
    """Composite importance score for an (uncovered) saint, plus the component
    parts used to render the report row."""
    traditions = len(split_multi(r["Tradition of Veneration"]))
    has_intercession = 1 if r["Commonly Asked Intercessions"].strip() else 0
    has_vocation = 1 if r["Vocation"].strip() else 0
    has_experience = 1 if r["Life Experience"].strip() else 0
    # Count fixed-date feasts; a movable-only feast still counts as one.
    feast_count = len(FEAST_RE.findall(r["Feast Day(s)"]) or [None])

    score = (2 * traditions + 3 * has_intercession + has_vocation
             + has_experience + feast_count)
    parts = {
        "traditions": traditions,
        "intercession": bool(has_intercession),
        "vocation": bool(has_vocation),
        "experience": bool(has_experience),
        "feasts": feast_count,
    }
    return score, parts


def priority_ranking(rows: list[dict[str, str]],
                     images: dict[str, dict[str, str]]
                     ) -> list[tuple[int, dict[str, str], dict[str, object]]]:
    """Saints WITHOUT a real icon, ranked by priority_score descending.
    Ties broken by Saint ID so the ordering is stable/reproducible."""
    covered = {sid for sid, img in images.items() if img.get("path")}
    ranked = []
    for r in rows:
        if r["Saint ID"].strip() in covered:
            continue
        score, parts = priority_score(r)
        ranked.append((score, r, parts))
    ranked.sort(key=lambda t: (-t[0], t[1]["Saint ID"]))
    return ranked


def report_priority(rows: list[dict[str, str]], top: int | None = None) -> None:
    """Print a ranked, human-readable table of icon-less saints (issue #83).
    `top` limits the rows shown (None = all)."""
    images = load_saint_images()
    ranked = priority_ranking(rows, images)
    shown = ranked if top is None else ranked[:top]

    covered = sum(1 for img in images.values() if img.get("path"))
    print(f"Icon-priority report - {len(ranked)} saints without a real icon "
          f"({covered} of {len(rows)} covered). "
          f"Showing top {len(shown)}.\n")
    header = (f"{'Rank':>4}  {'OS-ID':<8} {'Name':<40} {'Score':>5}  "
              f"{'Trad':>4}  {'Interc':<6}  {'Vocat':<5}  {'Exp':<3}  {'Feasts':>6}")
    print(header)
    print("-" * len(header))
    for i, (score, r, parts) in enumerate(shown, 1):
        name = r["Name"]
        if len(name) > 40:
            name = name[:37] + "..."
        print(f"{i:>4}  {r['Saint ID']:<8} {name:<40} {score:>5}  "
              f"{parts['traditions']:>4}  {'yes' if parts['intercession'] else 'no':<6}  "
              f"{'yes' if parts['vocation'] else 'no':<5}  "
              f"{'yes' if parts['experience'] else 'no':<3}  {parts['feasts']:>6}")


# --------------------------------------------------------------------------- #
# Derived link fields
# --------------------------------------------------------------------------- #
def derive_links(name: str, hymn: str, icon: str, video: str) -> tuple[str, str, str]:
    q = urllib.parse.quote_plus(name)
    hymn = hymn.strip() or f"https://www.google.com/search?q={q}+apolytikion+troparion"
    icon = icon.strip() or f"https://www.google.com/search?tbm=isch&q={q}+orthodox+icon"
    video = video.strip() or f"https://www.youtube.com/results?search_query={q}+orthodox+saint"
    return hymn, icon, video


def work_link(title: str, name: str) -> dict:
    """A Works-by/about entry rendered as {title, search-URL}. We never host the
    text; we link to a Google search for it (copyright-safe, §9)."""
    q = urllib.parse.quote_plus(f'"{title}" {name}')
    return {"t": title, "u": f"https://www.google.com/search?q={q}"}


def load_vendors() -> list[dict[str, str]]:
    """Icon vendors to link out to (data/vendors.csv: vendor,url_template).
    {q} in the template is replaced by the URL-encoded saint name. Links only —
    no vendor imagery is reproduced (§9); images await an affiliate agreement."""
    if not VENDORS_CSV.exists():
        return []
    out = []
    with VENDORS_CSV.open(encoding="utf-8", newline="") as f:
        for row in csv.DictReader(f):
            if row.get("vendor", "").strip() and row.get("url_template", "").strip():
                out.append({"vendor": row["vendor"].strip(),
                            "url_template": row["url_template"].strip()})
    return out


def vendor_links(name: str, vendors: list[dict[str, str]]) -> list[dict]:
    q = urllib.parse.quote_plus(name)
    return [{"vendor": v["vendor"], "url": v["url_template"].replace("{q}", q)}
            for v in vendors]


# --------------------------------------------------------------------------- #
# Self-hosted saint portraits (data/saint_images.csv). One row per saint, keyed
# by Saint ID, pointing at an image under static/ (Astro publicDir). The build
# joins these into the record as `image` (+ credit/license/source) and the
# frontend's tiered SaintAvatar shows the real icon instead of the monogram.
# Licensing is enforced here so an unlicensed image can never deploy (§9).
# --------------------------------------------------------------------------- #
def load_saint_images() -> dict[str, dict[str, str]]:
    """saint_id -> {path, license, credit, source}. Empty if the file is absent."""
    out: dict[str, dict[str, str]] = {}
    if not SAINT_IMAGES_CSV.exists():
        return out
    with SAINT_IMAGES_CSV.open(encoding="utf-8", newline="") as f:
        reader = csv.DictReader(f)
        if reader.fieldnames != SAINT_IMAGES_HEADER:
            sys.exit(f"FATAL: {SAINT_IMAGES_CSV} header must be "
                     f"{SAINT_IMAGES_HEADER}, got {reader.fieldnames!r}")
        for row in reader:
            sid = (row.get("saint_id") or "").strip()
            if not sid:
                continue
            out[sid] = {
                "path": (row.get("image_path") or "").strip(),
                "license": (row.get("license") or "").strip(),
                "credit": (row.get("credit") or "").strip(),
                "source": (row.get("source") or "").strip(),
            }
    return out


def image_thumb(path: str) -> str | None:
    """static/-relative avatar thumb for a self-hosted portrait, or None.
    Thumbs mirror static/icons/ under static/icons/thumbs/ as JPEGs
    (scripts/make_icon_thumbs.py; the download pipeline emits them on ingest).
    Returned only when the file actually exists, so a missing thumb degrades
    to the full-size portrait — never a broken image. Absolute URLs and
    non-icons/ paths have no thumb."""
    if re.match(r"^(https?:)?//", path) or not path.startswith("icons/"):
        return None
    rel = path[len("icons/"):]
    stem, _, _ = rel.rpartition(".")
    thumb = f"icons/thumbs/{stem or rel}.jpg"
    return thumb if (STATIC / thumb).is_file() else None


def validate_saint_images(valid_ids: set[str],
                          permissions: dict[str, dict[str, str]] | None = None
                          ) -> tuple[list[str], list[str]]:
    """Validate data/saint_images.csv against §9: known saint, an existing local
    file, and an accepted open license (with a credit when required) OR a
    Permission:<vendor> token validated against the image-permission registry."""
    errors: list[str] = []
    warnings: list[str] = []
    if not SAINT_IMAGES_CSV.exists():
        return errors, warnings
    if permissions is None:
        permissions = load_image_permissions()
    with SAINT_IMAGES_CSV.open(encoding="utf-8", newline="") as f:
        reader = csv.DictReader(f)
        if reader.fieldnames != SAINT_IMAGES_HEADER:
            return ([f"saint_images.csv header must be {SAINT_IMAGES_HEADER}, "
                     f"got {reader.fieldnames!r}"], warnings)
        seen: set[str] = set()
        for i, row in enumerate(reader, 2):
            if not any((v or "").strip() for v in row.values()):
                continue
            sid = (row.get("saint_id") or "").strip()
            path = (row.get("image_path") or "").strip()
            lic = (row.get("license") or "").strip()
            credit = (row.get("credit") or "").strip()
            source = (row.get("source") or "").strip()
            where = f"saint_images.csv line {i}"

            if not sid:
                errors.append(f"{where}: empty saint_id.")
            elif not ID_RE.match(sid):
                errors.append(f"{where}: saint_id {sid!r} is not an OS-#### id.")
            elif sid not in valid_ids:
                errors.append(f"{where}: saint_id {sid!r} matches no saint.")
            elif sid in seen:
                errors.append(f"{where}: duplicate image row for {sid} "
                              "(one portrait per saint).")
            seen.add(sid)

            if not path:
                errors.append(f"{where} ({sid}): empty image_path.")
            elif not (STATIC / path).is_file():
                errors.append(f"{where} ({sid}): image_path {path!r} not found "
                              f"under static/ (expected {(STATIC / path)}).")
            elif image_thumb(path) is None:
                warnings.append(f"{where} ({sid}): no avatar thumb for {path!r} — "
                                "cards/finder will load the full-size portrait. "
                                "Run: python scripts/make_icon_thumbs.py")

            lic_errors, lic_warnings = validate_image_license(
                where, sid, lic, credit, source, permissions,
                subject="Self-hosted images", noun="image")
            errors.extend(lic_errors)
            warnings.extend(lic_warnings)
    return errors, warnings


def load_saint_depictions() -> dict[str, list[dict[str, str]]]:
    """saint_id -> ordered list of {path, license, credit, source, kind, tag,
    title, era, by} (the carousel cards). Empty if the file is absent. Unlike
    saint_images this is intentionally MANY rows per saint, kept in file order."""
    out: dict[str, list[dict[str, str]]] = {}
    if not SAINT_DEPICTIONS_CSV.exists():
        return out
    with SAINT_DEPICTIONS_CSV.open(encoding="utf-8", newline="") as f:
        reader = csv.DictReader(f)
        if reader.fieldnames != SAINT_DEPICTIONS_HEADER:
            sys.exit(f"FATAL: {SAINT_DEPICTIONS_CSV} header must be "
                     f"{SAINT_DEPICTIONS_HEADER}, got {reader.fieldnames!r}")
        for row in reader:
            sid = (row.get("saint_id") or "").strip()
            if not sid:
                continue
            out.setdefault(sid, []).append({
                "path": (row.get("image_path") or "").strip(),
                "license": (row.get("license") or "").strip(),
                "credit": (row.get("credit") or "").strip(),
                "source": (row.get("source") or "").strip(),
                "kind": (row.get("kind") or "").strip(),
                "tag": (row.get("tag") or "").strip(),
                "title": (row.get("title") or "").strip(),
                "era": (row.get("era") or "").strip(),
                "by": (row.get("by") or "").strip(),
            })
    return out


def validate_saint_depictions(valid_ids: set[str],
                              permissions: dict[str, dict[str, str]] | None = None
                              ) -> tuple[list[str], list[str]]:
    """Validate data/saint_depictions.csv: known saint, an existing local file, a
    title, a known kind, and the SAME license gate as saint_images (open license
    with a credit when required, OR a Permission:<vendor> token validated against
    the registry — revoked vendors warn and are excluded). Many rows per saint are
    expected; only an exact (saint_id, image_path) repeat is a duplicate."""
    errors: list[str] = []
    warnings: list[str] = []
    if not SAINT_DEPICTIONS_CSV.exists():
        return errors, warnings
    if permissions is None:
        permissions = load_image_permissions()
    with SAINT_DEPICTIONS_CSV.open(encoding="utf-8", newline="") as f:
        reader = csv.DictReader(f)
        if reader.fieldnames != SAINT_DEPICTIONS_HEADER:
            return ([f"saint_depictions.csv header must be {SAINT_DEPICTIONS_HEADER}, "
                     f"got {reader.fieldnames!r}"], warnings)
        seen: set[tuple[str, str]] = set()
        for i, row in enumerate(reader, 2):
            if not any((v or "").strip() for v in row.values()):
                continue
            sid = (row.get("saint_id") or "").strip()
            path = (row.get("image_path") or "").strip()
            lic = (row.get("license") or "").strip()
            credit = (row.get("credit") or "").strip()
            source = (row.get("source") or "").strip()
            kind = (row.get("kind") or "").strip()
            title = (row.get("title") or "").strip()
            where = f"saint_depictions.csv line {i}"

            if not sid:
                errors.append(f"{where}: empty saint_id.")
            elif not ID_RE.match(sid):
                errors.append(f"{where}: saint_id {sid!r} is not an OS-#### id.")
            elif sid not in valid_ids:
                errors.append(f"{where}: saint_id {sid!r} matches no saint.")
            elif (sid, path) in seen:
                errors.append(f"{where}: duplicate depiction row for {sid} "
                              f"and image {path!r}.")
            seen.add((sid, path))

            if not path:
                errors.append(f"{where} ({sid}): empty image_path.")
            elif not (STATIC / path).is_file():
                errors.append(f"{where} ({sid}): image_path {path!r} not found "
                              f"under static/ (expected {(STATIC / path)}).")

            if not title:
                errors.append(f"{where} ({sid}): empty title (the card heading).")
            if kind and kind not in DEPICTION_KINDS:
                errors.append(f"{where} ({sid}): kind {kind!r} is not one of "
                              f"{sorted(DEPICTION_KINDS)}.")

            lic_errors, lic_warnings = validate_image_license(
                where, sid, lic, credit, source, permissions,
                subject="Depictions", noun="depiction")
            errors.extend(lic_errors)
            warnings.extend(lic_warnings)
    return errors, warnings


def load_image_permissions() -> dict[str, dict[str, str]]:
    """vendor_slug -> {name, attribution, homepage, granted, status, terms}.
    Empty if the file is absent. Loads ALL rows (incl. revoked) so callers can
    decide; validation enforces correctness separately."""
    out: dict[str, dict[str, str]] = {}
    if not IMAGE_PERMISSIONS_CSV.exists():
        return out
    with IMAGE_PERMISSIONS_CSV.open(encoding="utf-8", newline="") as f:
        reader = csv.DictReader(f)
        if reader.fieldnames != IMAGE_PERMISSIONS_HEADER:
            sys.exit(f"FATAL: {IMAGE_PERMISSIONS_CSV} header must be "
                     f"{IMAGE_PERMISSIONS_HEADER}, got {reader.fieldnames!r}")
        for row in reader:
            slug = (row.get("vendor_slug") or "").strip()
            if not slug:
                continue
            out[slug] = {
                "name": (row.get("vendor_name") or "").strip(),
                "attribution": (row.get("attribution") or "").strip(),
                "homepage": (row.get("homepage") or "").strip(),
                "granted": (row.get("granted") or "").strip(),
                "status": (row.get("status") or "").strip(),
                "terms": (row.get("terms") or "").strip(),
            }
    return out


def validate_image_permissions() -> tuple[list[str], list[str]]:
    """Validate data/image_permissions.csv: valid slug, known status, a name and
    attribution, and no duplicate slugs."""
    errors: list[str] = []
    warnings: list[str] = []
    if not IMAGE_PERMISSIONS_CSV.exists():
        return errors, warnings
    with IMAGE_PERMISSIONS_CSV.open(encoding="utf-8", newline="") as f:
        reader = csv.DictReader(f)
        if reader.fieldnames != IMAGE_PERMISSIONS_HEADER:
            return ([f"image_permissions.csv header must be "
                     f"{IMAGE_PERMISSIONS_HEADER}, got {reader.fieldnames!r}"], warnings)
        seen: set[str] = set()
        for i, row in enumerate(reader, 2):
            if not any((v or "").strip() for v in row.values()):
                continue
            slug = (row.get("vendor_slug") or "").strip()
            status = (row.get("status") or "").strip()
            where = f"image_permissions.csv line {i}"
            if not slug:
                errors.append(f"{where}: empty vendor_slug.")
            elif not SLUG_RE.match(slug):
                errors.append(f"{where}: vendor_slug {slug!r} is not kebab-case.")
            elif slug in seen:
                errors.append(f"{where}: duplicate vendor_slug {slug!r}.")
            else:
                seen.add(slug)
            if not (row.get("vendor_name") or "").strip():
                errors.append(f"{where} ({slug}): empty vendor_name.")
            if not (row.get("attribution") or "").strip():
                errors.append(f"{where} ({slug}): empty attribution.")
            if status not in PERMISSION_STATUSES:
                errors.append(f"{where} ({slug}): status {status!r} must be one of "
                              f"{sorted(PERMISSION_STATUSES)}.")
    return errors, warnings


# --------------------------------------------------------------------------- #
# Saint quotes (data/saint_quotes.csv). One verified quote per saint, keyed by
# Saint ID. The build joins the quote into the record as `quote` and the saint
# detail page renders it with a citation linking the public-domain source. Every
# quote must be transcribed verbatim from a PD translation and the `translation`
# field must name that PD source — enforced here so a copyrighted translation can
# never deploy (§9). Saints without a quote simply render no quote block.
# --------------------------------------------------------------------------- #
def load_saint_quotes() -> dict[str, dict[str, str]]:
    """saint_id -> {quote, work, locus, translation, source}. Empty if absent."""
    out: dict[str, dict[str, str]] = {}
    if not SAINT_QUOTES_CSV.exists():
        return out
    with SAINT_QUOTES_CSV.open(encoding="utf-8", newline="") as f:
        reader = csv.DictReader(f)
        if reader.fieldnames != SAINT_QUOTES_HEADER:
            sys.exit(f"FATAL: {SAINT_QUOTES_CSV} header must be "
                     f"{SAINT_QUOTES_HEADER}, got {reader.fieldnames!r}")
        for row in reader:
            sid = (row.get("saint_id") or "").strip()
            if not sid:
                continue
            out[sid] = {
                "quote": (row.get("quote") or "").strip(),
                "work": (row.get("work") or "").strip(),
                "locus": (row.get("locus") or "").strip(),
                "translation": (row.get("translation") or "").strip(),
                "source": (row.get("source_url") or "").strip(),
            }
    return out


PROFILE_FILE_RE = re.compile(r"^(OS-\d{4,})\.yaml$")
PROFILE_ID_RE = re.compile(r"^id:\s*(OS-\d{4,})\s*$", re.M)


def validate_saint_profiles(valid_ids: set[str]) -> tuple[list[str], list[str]]:
    """Cross-check src/content/profiles/*.yaml against the saints: filename is
    OS-####.yaml, names a real saint, and the file's `id:` matches the filename.
    Shape validation is Zod's job at astro build; this is the Python data gate.
    Empty/missing dir is allowed (no profiles yet)."""
    errors: list[str] = []
    warnings: list[str] = []
    if not PROFILES_DIR.is_dir():
        return errors, warnings
    for path in sorted(PROFILES_DIR.glob("*.yaml")):
        m = PROFILE_FILE_RE.match(path.name)
        if not m:
            errors.append(f"profiles/{path.name}: name must be OS-####.yaml")
            continue
        sid = m.group(1)
        if sid not in valid_ids:
            errors.append(f"profiles/{path.name}: {sid} is not a known Saint ID")
        body_id = PROFILE_ID_RE.search(path.read_text(encoding="utf-8"))
        if not body_id:
            errors.append(f"profiles/{path.name}: missing an `id:` field")
        elif body_id.group(1) != sid:
            errors.append(
                f"profiles/{path.name}: id {body_id.group(1)} != filename {sid}"
            )
    return errors, warnings


def validate_saint_quotes(valid_ids: set[str]) -> tuple[list[str], list[str]]:
    """Validate data/saint_quotes.csv: known saint, one quote per saint, the quote
    and a citing work/source present, and a public-domain translation (§9)."""
    errors: list[str] = []
    warnings: list[str] = []
    if not SAINT_QUOTES_CSV.exists():
        return errors, warnings
    with SAINT_QUOTES_CSV.open(encoding="utf-8", newline="") as f:
        reader = csv.DictReader(f)
        if reader.fieldnames != SAINT_QUOTES_HEADER:
            return ([f"saint_quotes.csv header must be {SAINT_QUOTES_HEADER}, "
                     f"got {reader.fieldnames!r}"], warnings)
        seen: set[str] = set()
        for i, row in enumerate(reader, 2):
            if not any((v or "").strip() for v in row.values()):
                continue
            sid = (row.get("saint_id") or "").strip()
            quote = (row.get("quote") or "").strip()
            work = (row.get("work") or "").strip()
            translation = (row.get("translation") or "").strip()
            source = (row.get("source_url") or "").strip()
            where = f"saint_quotes.csv line {i}"

            if not sid:
                errors.append(f"{where}: empty saint_id.")
            elif not ID_RE.match(sid):
                errors.append(f"{where}: saint_id {sid!r} is not an OS-#### id.")
            elif sid not in valid_ids:
                errors.append(f"{where}: saint_id {sid!r} matches no saint.")
            elif sid in seen:
                errors.append(f"{where}: duplicate quote row for {sid} "
                              "(one quote per saint).")
            seen.add(sid)

            if not quote:
                errors.append(f"{where} ({sid}): empty quote.")
            if not translation:
                errors.append(f"{where} ({sid}): empty translation. A quote must "
                              "name its public-domain translation (§9).")
            elif not translation_ok(translation):
                errors.append(f"{where} ({sid}): translation {translation!r} does not "
                              "name an accepted public-domain source (ANF / NPNF / "
                              "(PD) / CC0). Copyrighted translations must not be "
                              "reproduced (§9) — link out instead.")
            if not source:
                errors.append(f"{where} ({sid}): empty source_url — a quote must be "
                              "verifiable against its public-domain source.")
            elif not source.lower().startswith(("http://", "https://")):
                warnings.append(f"{where} ({sid}): source_url {source!r} is not a URL.")
            if not work:
                warnings.append(f"{where} ({sid}): no 'work' cited for the quote.")
            if len(quote) > 500:
                warnings.append(f"{where} ({sid}): quote is {len(quote)} chars — long "
                                "for a 'famous quote' (consider trimming).")
    return errors, warnings


# --------------------------------------------------------------------------- #
# Group taxonomy (data/groups.csv + data/saint_groups.csv). Groups re-link the
# members of a collective commemoration; the join references any saint row
# (individual OR still-collective), so the taxonomy ships independently of the
# splitting backlog. Joined into each record as `groups` and emitted whole as
# public/groups.json (mirrors emit_themes_json) for the /group/<slug> pages.
# --------------------------------------------------------------------------- #
def load_groups() -> list[dict[str, str]]:
    """Ordered list of group definitions (by `sort` then `name`). Empty if absent."""
    out: list[dict[str, str]] = []
    if not GROUPS_CSV.exists():
        return out
    with GROUPS_CSV.open(encoding="utf-8", newline="") as f:
        reader = csv.DictReader(f)
        if reader.fieldnames != GROUPS_HEADER:
            sys.exit(f"FATAL: {GROUPS_CSV} header must be {GROUPS_HEADER}, "
                     f"got {reader.fieldnames!r}")
        for row in reader:
            slug = (row.get("slug") or "").strip()
            if not slug:
                continue
            sort_raw = (row.get("sort") or "").strip()
            out.append({
                "slug": slug,
                "saint_id": (row.get("saint_id") or "").strip(),
                "name": (row.get("name") or "").strip(),
                "type": (row.get("type") or "").strip(),
                "description": (row.get("description") or "").strip(),
                "feast": (row.get("feast") or "").strip(),
                "sort": int(sort_raw) if sort_raw.lstrip("-").isdigit() else 0,
                "rule": (row.get("rule") or "").strip(),
            })
    out.sort(key=lambda g: (g["sort"], g["name"]))
    return out


# Open/DYNAMIC synaxes (e.g. the New Martyrs & Confessors of Russia, All Saints
# of a region) have no fixed roster — membership grows as saints are glorified.
# Such a group carries a `rule` in groups.csv instead of explicit saint_groups
# rows; build.py computes its members by matching every saint against the rule,
# so a newly-added qualifying saint joins automatically.
#
# Rule grammar: ` && `-joined conditions; each condition is `field:v1|v2|…` and
# matches when ANY value is a case-insensitive substring of the saint's mapped
# field. Conditions are ANDed. An unknown field never matches (fails validation).
_RULE_FIELDS = {
    "rank": "Rank / Type",
    "era": "Era",
    "century": "Century",
    "region": "Region of Origin",
    "tradition": "Tradition of Veneration",
    "vocation": "Vocation",
    "life": "Life Experience",
    "notes": "Notes",
    "name": "Name",
}


def _parse_rule(rule: str) -> list[tuple[str, list[str]]] | None:
    """Parse a membership rule into [(saints.csv column, [values]), …], or None
    if malformed (unknown field, missing `:`, or empty)."""
    conds: list[tuple[str, list[str]]] = []
    for raw in rule.split("&&"):
        cond = raw.strip()
        if not cond or ":" not in cond:
            return None
        key, vals = cond.split(":", 1)
        col = _RULE_FIELDS.get(key.strip().lower())
        values = [v.strip().lower() for v in vals.split("|") if v.strip()]
        if not col or not values:
            return None
        conds.append((col, values))
    return conds or None


def _saint_matches(saint: dict[str, str], conds: list[tuple[str, list[str]]]) -> bool:
    """True when the saint satisfies every parsed condition."""
    for col, values in conds:
        cell = (saint.get(col, "") or "").lower()
        if not any(v in cell for v in values):
            return False
    return True


def dynamic_group_members(
    groups: list[dict], rows: list[dict[str, str]]
) -> dict[str, list[str]]:
    """slug -> ordered [Saint ID] for each group carrying a `rule`, computed by
    matching every saint against the rule (sorted by Name). Malformed rules
    yield no members (validation reports them separately)."""
    out: dict[str, list[str]] = {}
    for g in groups:
        rule = g.get("rule", "")
        if not rule:
            continue
        conds = _parse_rule(rule)
        if not conds:
            continue
        matched = [r for r in rows if _saint_matches(r, conds)]
        matched.sort(key=lambda r: r["Name"].strip().lower())
        out[g["slug"]] = [r["Saint ID"].strip() for r in matched
                          if r["Saint ID"].strip()]
    return out


def group_members_detail(
    saints_by_name: dict[str, dict[str, str]] | None = None,
) -> dict[str, list[dict[str, str]]]:
    """group_slug -> ordered [{saint_id, name}], INCLUDING name-only members
    (blank saint_id, name carried in the `role` column). This is what the group
    saint-profile's "Members of this Group" section renders. `saints_by_name`
    maps Saint ID -> row so a member's display name is the saint's own Name;
    absent that, the `role` label is used."""
    saints_by_name = saints_by_name or {}
    rows: dict[str, list[tuple[int, int, dict[str, str]]]] = {}
    if not SAINT_GROUPS_CSV.exists():
        return {}
    with SAINT_GROUPS_CSV.open(encoding="utf-8", newline="") as f:
        for i, row in enumerate(csv.DictReader(f)):
            slug = (row.get("group_slug") or "").strip()
            if not slug:
                continue
            sid = (row.get("saint_id") or "").strip()
            role = (row.get("role") or "").strip()
            order_raw = (row.get("order") or "").strip()
            order = int(order_raw) if order_raw.lstrip("-").isdigit() else 10**9
            saint = saints_by_name.get(sid)
            name = (saint["Name"].strip() if saint else "") or role
            member = {"saint_id": sid, "name": name}
            if role and role != name:
                member["role"] = role
            rows.setdefault(slug, []).append((order, i, member))
    return {slug: [m for _o, _i, m in sorted(items, key=lambda t: (t[0], t[1]))]
            for slug, items in rows.items()}


def load_saint_groups() -> dict[str, list[str]]:
    """saint_id -> [group_slug, …] (membership), preserving group-then-row order."""
    out: dict[str, list[str]] = {}
    if not SAINT_GROUPS_CSV.exists():
        return out
    with SAINT_GROUPS_CSV.open(encoding="utf-8", newline="") as f:
        reader = csv.DictReader(f)
        if reader.fieldnames != SAINT_GROUPS_HEADER:
            sys.exit(f"FATAL: {SAINT_GROUPS_CSV} header must be "
                     f"{SAINT_GROUPS_HEADER}, got {reader.fieldnames!r}")
        for row in reader:
            sid = (row.get("saint_id") or "").strip()
            slug = (row.get("group_slug") or "").strip()
            if not sid or not slug:
                continue
            out.setdefault(sid, []).append(slug)
    return out


def group_members() -> dict[str, list[str]]:
    """group_slug -> [saint_id, …] member list (file order)."""
    out: dict[str, list[str]] = {}
    if not SAINT_GROUPS_CSV.exists():
        return out
    with SAINT_GROUPS_CSV.open(encoding="utf-8", newline="") as f:
        for row in csv.DictReader(f):
            sid = (row.get("saint_id") or "").strip()
            slug = (row.get("group_slug") or "").strip()
            if sid and slug:
                out.setdefault(slug, []).append(sid)
    return out


def validate_groups(valid_ids: set[str]) -> tuple[list[str], list[str]]:
    """Validate the group taxonomy join files: enumerated `type`, unique slugs,
    every membership reference resolves (group + saint), no duplicate membership."""
    errors: list[str] = []
    warnings: list[str] = []
    if not GROUPS_CSV.exists() and not SAINT_GROUPS_CSV.exists():
        return errors, warnings

    # A group's own saint_id (the OS-#### its /saint/<id> page is served at) must
    # be unique and must NOT collide with an active saint or a retired tombstone.
    retired: set[str] = set()
    if RETIRED_IDS_CSV.exists():
        with RETIRED_IDS_CSV.open(encoding="utf-8", newline="") as f:
            retired = {(r.get("retired_id") or "").strip()
                       for r in csv.DictReader(f)}
    slugs: set[str] = set()
    group_saint_ids: set[str] = set()
    if GROUPS_CSV.exists():
        with GROUPS_CSV.open(encoding="utf-8", newline="") as f:
            reader = csv.DictReader(f)
            if reader.fieldnames != GROUPS_HEADER:
                return ([f"groups.csv header must be {GROUPS_HEADER}, "
                         f"got {reader.fieldnames!r}"], warnings)
            for i, row in enumerate(reader, 2):
                if not any((v or "").strip() for v in row.values()):
                    continue
                slug = (row.get("slug") or "").strip()
                gid = (row.get("saint_id") or "").strip()
                gtype = (row.get("type") or "").strip()
                sort_raw = (row.get("sort") or "").strip()
                where = f"groups.csv line {i}"
                if not slug:
                    errors.append(f"{where}: empty slug.")
                elif not SLUG_RE.match(slug):
                    errors.append(f"{where}: slug {slug!r} must be kebab-case "
                                  "([a-z0-9] words joined by hyphens).")
                elif slug in slugs:
                    errors.append(f"{where}: duplicate group slug {slug!r}.")
                slugs.add(slug)
                # A blank saint_id is fine before a build assigns it (§6); a
                # present one must be well-formed, unique, and un-collided.
                if gid:
                    if not ID_RE.match(gid):
                        errors.append(f"{where} ({slug}): saint_id {gid!r} is not "
                                      "an OS-#### id.")
                    elif gid in group_saint_ids:
                        errors.append(f"{where} ({slug}): duplicate group saint_id "
                                      f"{gid!r}.")
                    elif gid in valid_ids:
                        errors.append(f"{where} ({slug}): saint_id {gid!r} collides "
                                      "with a saint in saints.csv.")
                    elif gid in retired:
                        errors.append(f"{where} ({slug}): saint_id {gid!r} is a "
                                      "retired id and must not be reused.")
                    group_saint_ids.add(gid)
                if gtype not in GROUP_TYPES:
                    errors.append(f"{where} ({slug}): type {gtype!r} is not one of "
                                  f"{sorted(GROUP_TYPES)}.")
                if not (row.get("name") or "").strip():
                    errors.append(f"{where} ({slug}): empty name.")
                if sort_raw and not sort_raw.lstrip("-").isdigit():
                    errors.append(f"{where} ({slug}): sort {sort_raw!r} is not an integer.")
                # A dynamic/open group carries a membership `rule` (see
                # dynamic_group_members). It must parse and name known fields.
                rule = (row.get("rule") or "").strip()
                if rule and _parse_rule(rule) is None:
                    errors.append(
                        f"{where} ({slug}): membership rule {rule!r} is malformed "
                        f"(use `field:v1|v2 && …`, fields: {sorted(_RULE_FIELDS)}).")

    if SAINT_GROUPS_CSV.exists():
        with SAINT_GROUPS_CSV.open(encoding="utf-8", newline="") as f:
            reader = csv.DictReader(f)
            if reader.fieldnames != SAINT_GROUPS_HEADER:
                return ([f"saint_groups.csv header must be {SAINT_GROUPS_HEADER}, "
                         f"got {reader.fieldnames!r}"], warnings)
            pairs: set[tuple[str, str]] = set()
            for i, row in enumerate(reader, 2):
                if not any((v or "").strip() for v in row.values()):
                    continue
                slug = (row.get("group_slug") or "").strip()
                sid = (row.get("saint_id") or "").strip()
                role = (row.get("role") or "").strip()
                order_raw = (row.get("order") or "").strip()
                where = f"saint_groups.csv line {i}"
                if not slug:
                    errors.append(f"{where}: empty group_slug.")
                elif slug not in slugs:
                    errors.append(f"{where}: group_slug {slug!r} matches no group "
                                  "in groups.csv.")
                # A member may be name-only (blank saint_id + a `role` name) for a
                # saint who has no individual row yet; a present saint_id must
                # resolve to a real saint.
                if not sid:
                    if not role:
                        errors.append(f"{where}: a member needs a saint_id or, for "
                                      "a name-only member, a `role` name.")
                elif not ID_RE.match(sid):
                    errors.append(f"{where}: saint_id {sid!r} is not an OS-#### id.")
                elif sid not in valid_ids:
                    errors.append(f"{where}: saint_id {sid!r} matches no saint.")
                if slug and sid:
                    if (slug, sid) in pairs:
                        errors.append(f"{where}: duplicate membership "
                                      f"({slug}, {sid}).")
                    pairs.add((slug, sid))
                if order_raw and not order_raw.lstrip("-").isdigit():
                    errors.append(f"{where} ({slug}/{sid}): order {order_raw!r} "
                                  "is not an integer.")
    return errors, warnings


# --------------------------------------------------------------------------- #
# Retired IDs (data/retired_ids.csv): tombstone of every deduplicated saint row.
# Two invariants hold at all times: (1) a retired ID must NOT exist as an active
# Saint ID — retiring means deleting the row; (2) the canonical_id it points to
# MUST exist — the canonical row must not itself have been deleted.
# --------------------------------------------------------------------------- #
def validate_retired_ids(valid_ids: set[str]) -> tuple[list[str], list[str]]:
    """Enforce retired-IDs invariants:
      • retired_id must not appear in active data (the row was deleted).
      • canonical_id must appear in active data (the keeper still exists)."""
    errors: list[str] = []
    warnings: list[str] = []
    if not RETIRED_IDS_CSV.exists():
        return errors, warnings
    with RETIRED_IDS_CSV.open(encoding="utf-8", newline="") as f:
        reader = csv.DictReader(f)
        if reader.fieldnames != RETIRED_IDS_HEADER:
            return ([f"retired_ids.csv header must be {RETIRED_IDS_HEADER}, "
                     f"got {reader.fieldnames!r}"], warnings)
        for i, row in enumerate(reader, 2):
            if not any((v or "").strip() for v in row.values()):
                continue
            rid = (row.get("retired_id") or "").strip()
            cid = (row.get("canonical_id") or "").strip()
            where = f"retired_ids.csv line {i}"
            if rid and rid in valid_ids:
                errors.append(
                    f"{where}: retired ID {rid!r} still appears as an active "
                    "Saint ID — delete the row from saints.csv."
                )
            if cid and cid not in valid_ids:
                errors.append(
                    f"{where}: canonical ID {cid!r} is not present in "
                    "saints.csv — the canonical entry may have been removed."
                )
    return errors, warnings


# --------------------------------------------------------------------------- #
# Name variants (data/name_variants.csv): equivalence groups of given-name
# forms — English nicknames + cross-language/transliteration variants. The
# build expands each saint's search haystack with the other forms in any group
# their Name / Also Known As belongs to, so e.g. searching "Lucy" finds Lucia
# and "Ivan" finds John, without hand-editing every row.
# --------------------------------------------------------------------------- #
def fold(s: str) -> str:
    """Lowercase + strip diacritics, so 'Étienne' and 'etienne' compare equal."""
    nfkd = unicodedata.normalize("NFKD", s)
    return "".join(c for c in nfkd if not unicodedata.combining(c)).lower()


def name_tokens(text: str) -> set[str]:
    """Folded alphabetic word tokens of a name string."""
    return {fold(t) for t in re.findall(r"[^\W\d_]+", text, flags=re.UNICODE)}


def load_name_variants() -> dict[str, list[str]]:
    """form (folded) -> the full list of display forms in that form's group."""
    lookup: dict[str, list[str]] = {}
    if not NAME_VARIANTS_CSV.exists():
        return lookup
    with NAME_VARIANTS_CSV.open(encoding="utf-8", newline="") as f:
        reader = csv.reader(f)
        header = next(reader, None)
        if header != ["group", "names"]:
            sys.exit(f"FATAL: {NAME_VARIANTS_CSV} header must be 'group,names', got {header!r}")
        for row in reader:
            if not row or not any(c.strip() for c in row):
                continue
            forms = split_multi(row[1]) if len(row) > 1 else []
            for form in forms:
                lookup[fold(form)] = forms
    return lookup


def validate_name_variants() -> list[str]:
    """A form must sit in exactly one group; each group needs >= 2 forms."""
    errs: list[str] = []
    if not NAME_VARIANTS_CSV.exists():
        return errs
    with NAME_VARIANTS_CSV.open(encoding="utf-8", newline="") as f:
        reader = csv.reader(f)
        header = next(reader, None)
        if header != ["group", "names"]:
            return [f"name_variants.csv header must be 'group,names', got {header!r}"]
        groups: set[str] = set()
        seen: dict[str, str] = {}
        for i, row in enumerate(reader, 2):
            if not row or not any(c.strip() for c in row):
                continue
            grp = row[0].strip()
            forms = split_multi(row[1]) if len(row) > 1 else []
            if not grp:
                errs.append(f"name_variants.csv line {i}: empty group key")
            elif grp in groups:
                errs.append(f"name_variants.csv: duplicate group '{grp}'")
            groups.add(grp)
            if len(forms) < 2:
                errs.append(f"name_variants.csv group '{grp}': needs at least 2 forms")
            for form in forms:
                key = fold(form)
                if key in seen and seen[key] != grp:
                    errs.append(f"name_variants.csv: form '{form}' is in both "
                                f"'{seen[key]}' and '{grp}' (a form must belong to one group)")
                seen[key] = grp
    return errs


def variant_forms(r: dict[str, str], lookup: dict[str, list[str]]) -> list[str]:
    """Extra display-name forms a saint can be found by, beyond what already
    appears in their Name / Also Known As."""
    if not lookup:
        return []
    present = name_tokens(r["Name"])
    for a in split_multi(r["Also Known As"]):
        present |= name_tokens(a)
    added: list[str] = []
    seen: set[str] = set()
    for tok in present:
        for form in lookup.get(tok, []):
            ff = fold(form)
            if ff in present or ff in seen:
                continue
            seen.add(ff)
            added.append(form)
    return sorted(added)


# --------------------------------------------------------------------------- #
# Emit data.json
# --------------------------------------------------------------------------- #
def to_record(r: dict[str, str], vendors: list[dict[str, str]] | None = None,
              name_variants: dict[str, list[str]] | None = None,
              images: dict[str, dict[str, str]] | None = None,
              quotes: dict[str, dict[str, str]] | None = None,
              saint_groups: dict[str, list[str]] | None = None,
              groups_by_slug: dict[str, dict] | None = None,
              permissions: dict[str, dict[str, str]] | None = None,
              depictions: dict[str, list[dict[str, str]]] | None = None) -> dict:
    """Transform one saints.csv row into one public/data.json record.

    Joins in everything keyed by the saint's ID — portrait (+thumb), quote,
    groups, depiction cards, curated/derived links, computed themes — splits
    multi-value cells into arrays, and builds the `search` haystack (name +
    facets + name variants) the finder matches against. The optional params
    are the pre-loaded join tables; emit_data_json() passes them once so the
    CSVs aren't re-read per row (None = load on demand, used by unit tests).
    Images from a revoked permission vendor are dropped here (monogram
    fallback); permission images carry imagePermission/imageVendor fields,
    open-license images carry imageLicense/imageCredit."""
    if vendors is None:
        vendors = load_vendors()
    if name_variants is None:
        name_variants = load_name_variants()
    if images is None:
        images = load_saint_images()
    if quotes is None:
        quotes = load_saint_quotes()
    if saint_groups is None:
        saint_groups = load_saint_groups()
    if groups_by_slug is None:
        groups_by_slug = {g["slug"]: g for g in load_groups()}
    if permissions is None:
        permissions = load_image_permissions()
    if depictions is None:
        depictions = load_saint_depictions()
    rec: dict = {}
    for col, key in JSON_KEYS.items():
        val = r[col]
        if key in ARRAY_KEYS:
            rec[key] = split_multi(val)
        else:
            rec[key] = val.strip()
    rec["months"] = parse_months(r["Feast Day(s)"])
    rec["feastSort"] = feast_sort(r["Feast Day(s)"])
    rec["hymn"], rec["icon"], rec["video"] = derive_links(
        r["Name"], r["Hymn / Apolytikion"], r["Icon"], r["Video / Media"]
    )
    # Works by/about -> {title, search-URL}; plus per-saint icon-vendor links.
    rec["works"] = [work_link(t, r["Name"]) for t in rec["works"]]
    rec["about"] = [work_link(t, r["Name"]) for t in rec["about"]]
    rec["vendors"] = vendor_links(r["Name"], vendors)
    # Self-hosted real portrait (data/saint_images.csv), if one exists for this
    # saint. `image` is a static/-relative path the frontend base-prefixes; the
    # tiered SaintAvatar then shows it instead of the monogram. Attribution
    # (credit/license/source) rides along for the detail-page caption.
    img = images.get(r["Saint ID"].strip())
    if img and img.get("path"):
        slug = permission_slug(img.get("license", ""))
        if slug is not None:
            vendor = permissions.get(slug)
            # Publish a permission image only if the vendor grant is active.
            if vendor and vendor.get("status") != "revoked":
                rec["image"] = img["path"]
                thumb = image_thumb(img["path"])
                if thumb:
                    rec["imageThumb"] = thumb
                rec["imagePermission"] = True
                rec["imageVendor"] = vendor.get("name", "")
                rec["imageAttribution"] = vendor.get("attribution", "")
                rec["imageVendorHome"] = vendor.get("homepage", "")
                if img.get("source"):
                    rec["imageSource"] = img["source"]
            # revoked / unknown vendor -> no image key (monogram fallback)
        else:
            rec["image"] = img["path"]
            thumb = image_thumb(img["path"])
            if thumb:
                rec["imageThumb"] = thumb
            if img.get("license"):
                rec["imageLicense"] = img["license"]
            if img.get("credit"):
                rec["imageCredit"] = img["credit"]
            if img.get("source"):
                rec["imageSource"] = img["source"]
    # Additional depictions (data/saint_depictions.csv) — the saint page's
    # "Depictions & Icons" carousel. Many per saint, in file order. Each carries
    # the card presentation (kind/tag/title/era/by) plus its rights: a permission
    # depiction (active vendor) gets permission/vendor/attribution; an open-license
    # one keeps license/credit. A revoked vendor's depiction is dropped (like the
    # hero image). `source` is the per-card outbound link (grant condition for
    # permission cards: each links to its specific icon page).
    deps = depictions.get(r["Saint ID"].strip())
    if deps:
        cards: list[dict] = []
        for d in deps:
            if not d.get("path"):
                continue
            card: dict = {
                "image": d["path"],
                "kind": d.get("kind") or "museum",
                "title": d.get("title", ""),
            }
            for k in ("tag", "era", "by"):
                if d.get(k):
                    card[k] = d[k]
            if d.get("source"):
                card["source"] = d["source"]
            slug = permission_slug(d.get("license", ""))
            if slug is not None:
                vendor = permissions.get(slug)
                if not vendor or vendor.get("status") == "revoked":
                    continue  # unknown / revoked vendor → exclude the depiction
                card["permission"] = True
                card["vendor"] = vendor.get("name", "")
                card["attribution"] = vendor.get("attribution", "")
            else:
                if d.get("license"):
                    card["license"] = d["license"]
                if d.get("credit"):
                    card["credit"] = d["credit"]
            cards.append(card)
        if cards:
            rec["depictions"] = cards
    # Verified public-domain quote (data/saint_quotes.csv), if one exists. The
    # detail page renders `quote` with a citation; `quoteSource` links the PD
    # source so the wording is verifiable (§9). Saints without one render nothing.
    q = quotes.get(r["Saint ID"].strip())
    if q and q.get("quote"):
        rec["quote"] = q["quote"]
        if q.get("work"):
            rec["quoteWork"] = q["work"]
        if q.get("locus"):
            rec["quoteLocus"] = q["locus"]
        if q.get("translation"):
            rec["quoteTranslation"] = q["translation"]
        if q.get("source"):
            rec["quoteSource"] = q["source"]
    # Search haystack: name + aka + brief + notes + customs + all facet values.
    facets = []
    for col in CONTROLLED + FREE_MULTI + ["Brief Life", "Notes", "Customs & Traditions"]:
        facets.append(r[col])
    rec["search"] = " ".join(p for p in facets + [r["Name"]] if p).strip()
    # Name-variant expansion: make the saint findable by nickname / other-language
    # forms. The display forms power a "matched via" hint; the folded forms keep
    # the (accent-naive, lowercased) client substring search working for them.
    added = variant_forms(r, name_variants)
    if added:
        rec["variants"] = added
        extra = []
        for form in added:
            extra.append(form)
            folded = fold(form)
            if folded != form.lower():
                extra.append(folded)
        rec["search"] = (rec["search"] + " " + " ".join(extra)).strip()
    # Group taxonomy memberships (data/saint_groups.csv): each as {slug,name,type}
    # for the saint-page "Member of" links, ordered by the group's `sort`. Names
    # also feed the finder's Group facet (added to the search haystack).
    memberships = []
    for slug in saint_groups.get(r["Saint ID"].strip(), []):
        g = groups_by_slug.get(slug)
        if g:
            memberships.append({"slug": slug, "id": g.get("saint_id", ""),
                                "name": g["name"], "type": g["type"]})
    memberships.sort(key=lambda m: (groups_by_slug[m["slug"]]["sort"], m["name"]))
    if memberships:
        rec["groups"] = memberships
        # `groupNames` (plain string list) is what the finder's Group facet keys
        # on (valuesOf/facetCounts expect a string[] field named like the facet);
        # `groups` (objects) drives the saint-page "Commemorated With" slug links.
        rec["groupNames"] = [m["name"] for m in memberships]
        rec["search"] = (rec["search"] + " "
                         + " ".join(m["name"] for m in memberships)).strip()
    rec["themes"] = themes_mod.compute_themes(rec, r.get("Themes", ""))
    if rec["themes"]:
        label_words = " ".join(themes_mod.THEME_LABELS[s] for s in rec["themes"]
                               if s in themes_mod.THEME_LABELS)
        rec["search"] = (rec["search"] + " " + label_words).strip()
    return rec


def group_record(g: dict, members: list[dict[str, str]]) -> dict:
    """A group rendered as a saint-shaped record for public/data.json, so it
    flows into /saint/<id>, the finder, the calendar and the sitemap unchanged
    (only the quiz filters it out). `profile_type: "group"` steers the frontend
    to the GroupSaintProfile layout instead of SaintView. Facet arrays stay
    empty — a group is not an intercessor, so it never pollutes a facet filter."""
    rec: dict = {}
    for key in JSON_KEYS.values():
        rec[key] = [] if key in ARRAY_KEYS else ""
    rec["id"] = g["saint_id"]
    rec["name"] = g["name"]
    rec["feast"] = g["feast"]
    rec["brief"] = g["description"]
    rec["months"] = parse_months(g["feast"])
    rec["feastSort"] = feast_sort(g["feast"])
    rec["themes"] = []
    rec["vendors"] = []
    rec["profile_type"] = "group"
    rec["groupSlug"] = g["slug"]
    rec["groupType"] = g["type"]
    rec["members"] = members
    # Text-searchable by name, description and member names (finder search yes,
    # quiz no — the quiz island drops profile_type:"group").
    member_names = " ".join(m["name"] for m in members if m.get("name"))
    rec["search"] = " ".join(p for p in (g["name"], g["description"],
                                         member_names) if p).strip()
    return rec


def emit_data_json(rows: list[dict[str, str]]) -> list[dict]:
    """Write public/data.json: every saint as a record, sorted by feastSort.

    The Astro build's primary input (src/lib/data.ts reads it from disk at
    build time). Minified separators; returns the records for reuse by
    emit_themes_json()."""
    vendors = load_vendors()
    name_variants = load_name_variants()
    images = load_saint_images()
    quotes = load_saint_quotes()
    saint_groups = load_saint_groups()
    groups = load_groups()
    groups_by_slug = {g["slug"]: g for g in groups}
    permissions = load_image_permissions()
    depictions = load_saint_depictions()
    # Dynamic (rule-based) group membership — merge into the reverse index BEFORE
    # building records so a member saint shows "Member of <open synaxis>".
    by_id = {r["Saint ID"].strip(): r for r in rows}
    dyn = dynamic_group_members(groups, rows)
    for slug, ids in dyn.items():
        for sid in ids:
            lst = saint_groups.setdefault(sid, [])
            if slug not in lst:
                lst.append(slug)
    records = [to_record(r, vendors, name_variants, images, quotes,
                         saint_groups, groups_by_slug, permissions, depictions)
               for r in rows]
    # Group saint-profiles: one record per group carrying its own OS-#### id.
    members_by_group = group_members_detail(by_id)
    # A rule-based group's roster IS the dynamic set (sorted by name).
    for slug, ids in dyn.items():
        members_by_group[slug] = [
            {"saint_id": sid, "name": by_id[sid]["Name"].strip()}
            for sid in ids if sid in by_id
        ]
    records.extend(group_record(g, members_by_group.get(g["slug"], []))
                   for g in groups if g.get("saint_id"))
    records.sort(key=lambda x: x["feastSort"])
    PUBLIC.mkdir(exist_ok=True)
    with open(PUBLIC / "data.json", "w", encoding="utf-8") as f:
        json.dump(records, f, ensure_ascii=False, separators=(",", ":"))
    return records


def emit_groups_json(rows: list[dict[str, str]] | None = None) -> list[dict]:
    """Write public/groups.json: each group with its full member id list, for the
    pre-rendered /group/<slug> pages (mirrors emit_themes_json). `rows` (the
    saints) lets rule-based open groups contribute their dynamic membership."""
    groups = load_groups()
    members = group_members()
    if rows is not None:
        for slug, ids in dynamic_group_members(groups, rows).items():
            members[slug] = ids
    catalog = [{**g, "members": members.get(g["slug"], [])} for g in groups]
    PUBLIC.mkdir(exist_ok=True)
    with open(PUBLIC / "groups.json", "w", encoding="utf-8") as f:
        json.dump(catalog, f, ensure_ascii=False, separators=(",", ":"))
    print(f"  wrote public/groups.json ({len(catalog)} groups, "
          f"{sum(len(c['members']) for c in catalog)} memberships)")
    return catalog


def emit_themes_json(records: list[dict]) -> None:
    catalog = themes_mod.theme_catalog(records)
    payload = {"themes": catalog, "aliases": themes_mod.ALIASES}
    with open(PUBLIC / "themes.json", "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, separators=(",", ":"))
    shown = sum(1 for c in catalog if c["count"] > 0)
    print(f"  wrote {PUBLIC.relative_to(ROOT)}/themes.json "
          f"({shown}/{len(catalog)} themes populated)")


# --------------------------------------------------------------------------- #
# Emit xlsx
# --------------------------------------------------------------------------- #
def emit_xlsx(header: list[str], rows: list[dict[str, str]],
              vocab: dict[str, set[str]],
              feasts_header: list[str] | None = None,
              feasts_rows: list[dict[str, str]] | None = None,
              hosts_header: list[str] | None = None,
              hosts_rows: list[dict[str, str]] | None = None):
    from openpyxl import Workbook
    from openpyxl.styles import Font

    DIST.mkdir(exist_ok=True)
    wb = Workbook()

    ws = wb.active
    ws.title = "Saints Database"
    ws.append(header)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for r in rows:
        ws.append([r[c] for c in header])
    ws.freeze_panes = "A2"

    if feasts_rows:
        fs = wb.create_sheet("Feasts & Fasts")
        fs.append(feasts_header)
        for cell in fs[1]:
            cell.font = Font(bold=True)
        for r in feasts_rows:
            fs.append([r[c] for c in feasts_header])
        fs.freeze_panes = "A2"

    if hosts_rows:
        hs = wb.create_sheet("Heavenly Hosts")
        hs.append(hosts_header)
        for cell in hs[1]:
            cell.font = Font(bold=True)
        for r in hosts_rows:
            hs.append([r[c] for c in hosts_header])
        hs.freeze_panes = "A2"

    vs = wb.create_sheet("Controlled Vocabulary")
    vs.append(["category", "term"])
    for cell in vs[1]:
        cell.font = Font(bold=True)
    for cat in sorted(vocab):
        for term in sorted(vocab[cat]):
            vs.append([cat, term])

    rm = wb.create_sheet("Read Me")
    notes = [
        "Orthodox Saints Database — Excel export",
        "",
        f"{len(rows)} saints. Generated from data/saints.csv (source of truth).",
        "",
        "This dataset is a work in progress and is NOT an authoritative or official",
        "liturgical resource. It awaits review by competent clergy and sources.",
        "",
        "Code: MIT License. Data: CC0 1.0 (public domain dedication).",
    ]
    for line in notes:
        rm.append([line])

    out = DIST / "Orthodox_Saints_Database.xlsx"
    wb.save(out)
    print(f"  wrote {out.relative_to(ROOT)}")


def emit_sqlite(conn: sqlite3.Connection):
    PUBLIC.mkdir(exist_ok=True)
    out = PUBLIC / "saints.sqlite"
    if out.exists():
        out.unlink()
    disk = sqlite3.connect(str(out))
    conn.backup(disk)
    disk.close()
    print(f"  wrote {out.relative_to(ROOT)}")


# --------------------------------------------------------------------------- #
# Main
# --------------------------------------------------------------------------- #
def main() -> int:
    ap = argparse.ArgumentParser(description="Build the Orthodox Saints Database.")
    ap.add_argument("--check-only", action="store_true",
                    help="validate only; no output files")
    ap.add_argument("--xlsx-only", action="store_true",
                    help="emit only the Excel export")
    ap.add_argument("--no-xlsx", action="store_true",
                    help="skip the Excel export (lets the full build run without openpyxl)")
    ap.add_argument("--sqlite", action="store_true",
                    help="also emit public/saints.sqlite")
    ap.add_argument("--report", action="store_true",
                    help="print a ranked table of saints lacking a real icon, then exit "
                         "(local authoring aid; no files written)")
    ap.add_argument("--top", type=int, default=50, metavar="N",
                    help="rows to show in --report (0 = all; default 50)")
    args = ap.parse_args()

    vocab = load_vocab()
    header, rows = load_saints()
    f_header, f_rows = feastlib.load_feasts()
    h_header, h_rows = hostlib.load_hosts()

    # The priority report is a read-only authoring aid: validate quietly so the
    # ranking reflects the committed data, but never write files or assign IDs.
    if args.report:
        errors, _ = validate(header, rows, vocab)
        if errors:
            print(f"WARNING: {len(errors)} validation error(s) in the source data; "
                  "the report below reflects the data as committed.", file=sys.stderr)
        report_priority(rows, top=None if args.top == 0 else args.top)
        return 0

    g_header, g_rows = load_group_rows()
    if not args.check_only:
        # Saints and groups draw from ONE OS-#### counter (§6): seed it above the
        # highest number anywhere — active saints, existing group ids, and
        # retired tombstones — so neither space can ever collide or reuse.
        seed = max(next_id_seed(rows), next_id_seed(g_rows, "saint_id"),
                   retired_id_seed())
        saints_assigned, seed = assign_ids(rows, seed=seed)
        if saints_assigned:
            write_saints(header, rows)
        if g_header:
            groups_assigned, seed = assign_ids(g_rows, "saint_id", seed=seed)
            if groups_assigned:
                write_groups(g_header, g_rows)
        if feastlib.assign_ids(f_rows):
            feastlib.write_feasts(f_header, f_rows)
        # Heavenly Hosts (HH-####) run an INDEPENDENT id counter, like FF-####.
        if hostlib.assign_ids(h_rows):
            hostlib.write_hosts(h_header, h_rows)

    conn = build_db(header, rows, vocab)
    errors, warnings = validate(header, rows, vocab)
    saint_ids = {r["Saint ID"].strip() for r in rows}
    f_errors, f_warnings = feastlib.validate(f_rows, vocab, saint_ids)
    errors.extend(f_errors)
    warnings.extend(f_warnings)
    feast_ids = {r["Feast ID"].strip() for r in f_rows}
    h_errors, h_warnings = hostlib.validate(h_rows, vocab, saint_ids, feast_ids)
    errors.extend(h_errors)
    warnings.extend(h_warnings)

    # Finder-coverage nudges are bulk and low-signal; summarize them. Other
    # warnings (e.g. possible duplicate saints) are surfaced individually.
    coverage = [w for w in warnings if "finder coverage" in w]
    other = [w for w in warnings if "finder coverage" not in w]
    for w in other:
        print(f"WARN: {w}")
    if coverage:
        print(f"WARN: {len(coverage)} saint(s) without Commonly Asked "
              f"Intercessions (finder-coverage nudge; not a failure).")
    if errors:
        print(f"\nVALIDATION FAILED — {len(errors)} error(s):", file=sys.stderr)
        for e in errors:
            print(f"  ERROR: {e}", file=sys.stderr)
        return 1

    print(f"VALIDATION CLEAN — {len(rows)} saints, {len(f_rows)} feasts, "
          f"{len(h_rows)} hosts, {len(warnings)} warning(s), 0 errors.")

    report_coverage(rows)

    if args.check_only:
        return 0

    if args.xlsx_only:
        emit_xlsx(header, rows, vocab, f_header, f_rows, h_header, h_rows)
        return 0

    records = emit_data_json(rows)
    emit_themes_json(records)
    emit_groups_json(rows)
    feastlib.emit_feasts_json(f_rows)
    hostlib.emit_hosts_json(h_rows)
    print(f"  wrote public/data.json ({len(records)} records)")
    if not args.no_xlsx:
        emit_xlsx(header, rows, vocab, f_header, f_rows, h_header, h_rows)
    if args.sqlite:
        emit_sqlite(conn)

    conn.close()
    return 0


if __name__ == "__main__":
    sys.exit(main())
